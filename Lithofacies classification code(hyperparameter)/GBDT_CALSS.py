from sklearn.ensemble import GradientBoostingClassifier,AdaBoostClassifier
from sklearn.model_selection import GridSearchCV
from sklearn import preprocessing
import pandas as pd
import numpy as np
from sklearn.metrics import confusion_matrix,precision_score,recall_score,f1_score
import matplotlib.pyplot as plt
from sklearn.utils import shuffle
import openpyxl
import time
import matplotlib.pyplot as plt  #绘图库

random_seed = 42
train_ratio = 0.7
#读取数据并将数据转换为array格式
data_files = []
data_files = pd.read_csv("C:/Users/Administrator.DESKTOP-G2KEBOV/Desktop/岩相分类实验数据/4features949original data.csv")
data_files2 = pd.read_csv("C:/Users/Administrator.DESKTOP-G2KEBOV/Desktop/岩相分类实验数据/4features733new data.csv")

#设定为函数
def split_data(dataframe, train_ratio, random_state):
    '''
    将原始数据中每类数据都按照train_ratio的比例划分测试集和训练集
    参数
    dataframe：所读取的原始数据
    train_ratio:训练集所占比例
    random_state：打乱数据的随机数
    '''
    # 获取所有的类别
    lst = dataframe["class"].unique()
    # 创建固定维度的空数组以便拼接
    column_num = dataframe.shape[1]
    data_train = np.empty(shape=(0, column_num))
    data_test = np.empty(shape=(0, column_num))
    #循环对每类数据进行比例划分
    for i in lst:
        data_class = dataframe[dataframe["class"] == i]  # 筛选相同类的数据
        data_class = np.array(data_class)
        df = shuffle(data_class, random_state=random_state)  # 打乱次序
        train_size = int(len(df) * train_ratio)
        train = df[0:train_size, :]
        test = df[train_size:, :]
        data_train = np.concatenate((data_train, train), axis=0)
        data_test = np.concatenate((data_test, test), axis=0)
    return data_train, data_test
#绘图
labels_name = ['0','1','2','3','4','5','6','7','8','9','10','11','12','13']  #将横纵坐标标签集合赋值
def plot_confusion_matrix(cm,lables_name,title):
    cm = cm / cm.sum(axis=1)[:,np.newaxis]  #归一化
    plt.imshow(cm,interpolation='nearest')  #在特定的窗口上显示图像
    #plt.title(title)  #图像标题
    plt.matshow(cm,cmap=plt.cm.Reds)  #颜色
    plt.colorbar()
    num_class = np.array(range(len(lables_name)))  #获取标签的间隔数
    plt.xticks(num_class,lables_name,rotation=90)  #将标签印在x轴坐标上
    plt.yticks(num_class, lables_name)  # 将标签印在y轴坐标上
    plt.ylabel('Ture label')
    plt.xlabel('Predicted label')
    plt.savefig(title + 'GBDT.tiff')


data_train, data_test = split_data(data_files,train_ratio=train_ratio,random_state=random_seed)
data_train2, data_test2 = split_data(data_files2,train_ratio=train_ratio,random_state=random_seed)

data_train= np.concatenate((data_train,data_train2),axis=0)
data_test= np.concatenate((data_test,data_test2),axis=0)

#划分训练集和测试集的X，Y
X_train = data_train[:,1:]
X_test = data_test[:,1:]
Y_train = data_train[:,0]
Y_test = data_test[:,0]

#数据归一化
preprocessor = preprocessing.MinMaxScaler()
preprocessor.fit(np.concatenate((X_train,X_train),axis=0))
X_train = preprocessor.transform(X_train)
X_test = preprocessor.transform(X_test)

estimator = GradientBoostingClassifier()

GBDTClassifier_model = GradientBoostingClassifier(n_estimators = 110,learning_rate= 0.55,max_depth=2,random_state=1)
GBDTClassifier_model.fit(X_train,Y_train)

data_files1 = pd.read_csv("C:/Users/Administrator.DESKTOP-G2KEBOV/Desktop/949original data.csv")
data_files1 = preprocessor.transform(data_files1)
Y_pred = GBDTClassifier_model.predict(data_files1)

print(Y_pred)
y_pred = pd.DataFrame(Y_pred)
write = pd.ExcelWriter('C:/Users/Administrator.DESKTOP-G2KEBOV/Desktop/论文/新建 XLSX 工作表.xlsx')
y_pred.to_excel(write)
write.save()
write.close()
# for i in range(20,150,10):
#     param_grid = {"n_estimators": [i],
#                   #"max_features": ['sqrt'],
#                   "learning_rate": [0.55],
#                   "max_depth": [2],
#                   "random_state": [1]
#                   }
#
#     GBDTClassifier_model = GridSearchCV(estimator, param_grid, scoring='accuracy', cv=10 ,verbose=1 ,n_jobs=-1)
#     start = time.time()
#     GBDTClassifier_model.fit(X_train,Y_train)
#     end = time.time()
#     print('最优参数：',GBDTClassifier_model.best_params_)
#     #最优参数：
#     model = GBDTClassifier_model.best_estimator_
#     #测试集结果
#     Y_pred = model.predict(X_test)
#     cfm = confusion_matrix(Y_test,Y_pred)
#     precision = precision_score(Y_test,Y_pred,average='micro')
#     recall = recall_score(Y_test,Y_pred,average='macro')
#     f1 = f1_score(Y_test,Y_pred, average='macro')
#     timesum = end - start
#     print("准确率:",precision)
#     print("召回率:",recall)
#     print("F1分数:",f1)
#     print("耗时：", timesum,"\n")
#
#     #绘图
#     figname = 'cf_' + f'n_estimator_{i}-precision_{float(precision):2f}'
#     plot_confusion_matrix(cfm, labels_name, figname)  # 调用函数
#
#     #to excel
#     dataresult = [i, precision, recall, f1, timesum]
#     dataex = openpyxl.load_workbook('666输出结果.xlsx')
#     table = dataex.active
#     nrows = table.max_row # 获得行数
#     for i in range(len(dataresult)):
#         table.cell(nrows + 1, i + 1).value = dataresult[i]
#     dataex.save('666输出结果.xlsx')
