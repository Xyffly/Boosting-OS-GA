import pandas as pd, numpy as np
from sklearn.model_selection import train_test_split, GridSearchCV
from sklearn import metrics
import catboost as cb
from sklearn.utils import shuffle
from sklearn import preprocessing
from sklearn.metrics import confusion_matrix,precision_score,recall_score,f1_score
import matplotlib.pyplot as plt
from lightgbm.sklearn import LGBMClassifier
import openpyxl
import time
import matplotlib.pyplot as plt  #绘图库
from bitstring import BitArray
from deap import base, creator, tools, algorithms
from scipy.stats import bernoulli

random_seed = 42
train_ratio = 0.7
#读取数据并将数据转换为array格式
data_files = []
data_files = pd.read_csv("C:/Users/Administrator.DESKTOP-G2KEBOV/Desktop/岩相分类实验数据/4features949original data.csv")
data_files2 = pd.read_csv("C:/Users/Administrator.DESKTOP-G2KEBOV/Desktop/岩相分类实验数据/4features733new data.csv")
#设定为函数
#设定为函数
def split_data(dataframe, train_ratio, random_state):
    '''
    将原始数据中每类数据都按照train_ratio的比例划分测试集和训练集
    参数
    dataframe：所读取的原始数据
    train_ratio:训练集所占比例
    random_state：打乱数据的随机数
    '''
    # 获取所有的类别
    lst = dataframe["class"].unique()
    # 创建固定维度的空数组以便拼接
    column_num = dataframe.shape[1]
    data_train = np.empty(shape=(0, column_num))
    data_test = np.empty(shape=(0, column_num))
    #循环对每类数据进行比例划分
    for i in lst:
        data_class = dataframe[dataframe["class"] == i]  # 筛选相同类的数据
        data_class = np.array(data_class)
        df = shuffle(data_class, random_state=random_state)  # 打乱次序
        train_size = int(len(df) * train_ratio)
        train = df[0:train_size, :]
        test = df[train_size:, :]
        data_train = np.concatenate((data_train, train), axis=0)
        data_test = np.concatenate((data_test, test), axis=0)
    return data_train, data_test
#绘图
labels_name = ['0','1','2','3','4','5','6','7','8','9','10','11','12','13']  #将横纵坐标标签集合赋值
def plot_confusion_matrix(cm,lables_name,title):
    cm = cm / cm.sum(axis=1)[:,np.newaxis]  #归一化
    plt.imshow(cm,interpolation='nearest')  #在特定的窗口上显示图像
    #plt.title(title)  #图像标题
    plt.matshow(cm,cmap=plt.cm.Reds)  #颜色
    plt.colorbar()
    num_class = np.array(range(lables_name))  #获取标签的间隔数
    plt.xticks(num_class,lables_name,rotation=90)  #将标签印在x轴坐标上
    plt.yticks(num_class, lables_name)  # 将标签印在y轴坐标上
    plt.ylabel('Ture label')
    plt.xlabel('Predicted label')
    plt.savefig(title + 'LightGBM.tiff')


data_train, data_test = split_data(data_files,train_ratio=train_ratio,random_state=random_seed)
data_train2, data_test2 = split_data(data_files2,train_ratio=train_ratio,random_state=random_seed)

data_train= np.concatenate((data_train,data_train2),axis=0)
data_test= np.concatenate((data_test,data_test2),axis=0)

#划分训练集和测试集的X，Y
X_train = data_train[:,1:]
X_test = data_test[:,1:]
Y_train = data_train[:,0]
Y_test = data_test[:,0]

#数据归一化
preprocessor = preprocessing.MinMaxScaler()
preprocessor.fit(np.concatenate((X_train,X_train),axis=0))
X_train = preprocessor.transform(X_train)
X_test = preprocessor.transform(X_test)

def train_evaluate(ga_individual_solution):
    #参数解码
    n_estimators = BitArray(ga_individual_solution[0:4]) #2 ** 4 = 16
    learning_rate = BitArray(ga_individual_solution[4:8])
    max_depth = BitArray(ga_individual_solution[8:12])

    n_estimators = (n_estimators.uint + 1) * 10 # (10, 170, 10)
    learning_rate = (learning_rate.uint + 1) * 0.05 #(1 * 0.5, 17 * 0.5)
    max_depth = max_depth.uint + 1
    print('n_estimators:', n_estimators, 'learning_rate:', learning_rate, 'max_depth:', max_depth)

    estimator = LGBMClassifier(n_estimators = n_estimators, learning_rate=learning_rate,max_depth=max_depth, random_state=42)
    estimator.fit(X_train, Y_train)
    Y_pred = estimator.predict(X_test)
    cfm = confusion_matrix(Y_test,Y_pred)
    precision = precision_score(Y_test,Y_pred,average='micro')
    recall = recall_score(Y_test,Y_pred,average='macro')
    f1 = f1_score(Y_test,Y_pred, average='macro')
    print("准确率:",precision)
    print("召回率:",recall)
    print("F1分数:",f1,"\n")
    #to excel
    dataresult = [n_estimators, learning_rate, max_depth, precision, recall, f1]
    dataex = openpyxl.load_workbook('GA-LightGBM输出结果.xlsx')
    table = dataex.active
    nrows = table.max_row # 获得行数
    for i in range(len(dataresult)):
        table.cell(nrows + 1, i + 1).value = dataresult[i]
    dataex.save('GA-LightGBM输出结果.xlsx')
    dataex.close()
    return precision,

population_size = 10 #初始种群数
num_generations = 10 #迭代次数
gene_length = 12  #个体基因长度

#Implementation of Genetic Algorithm using DEAP python library.

#Since we try to minimise the loss values, we use the negation of the root mean squared loss as fitness function.
creator.create('FitnessMax', base.Fitness, weights = (1.0,))
creator.create('Individual', list, fitness = creator.FitnessMax)

#initialize the variables as bernoilli random variables
toolbox = base.Toolbox()
toolbox.register('binary', bernoulli.rvs, 0.5)
toolbox.register('individual', tools.initRepeat, creator.Individual, toolbox.binary, n = gene_length)
toolbox.register('population', tools.initRepeat, list , toolbox.individual)

#Ordered cross-over used for mating
toolbox.register('mate', tools.cxOrdered)
#Shuffle mutation to reorder the chromosomes
toolbox.register('mutate', tools.mutShuffleIndexes, indpb = 0.6)
#use roulette wheel selection algorithm 锦标赛
toolbox.register('select', tools.selTournament, tournsize = 2)
#training function used for evaluating fitness of individual solution.
toolbox.register('evaluate', train_evaluate)

population = toolbox.population(n = population_size)
r = algorithms.eaSimple(population, toolbox, cxpb = 0.4, mutpb = 0.1, ngen = num_generations, verbose = False)
